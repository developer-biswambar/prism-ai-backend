# backend/app/routes/file_routes.py - Enhanced with Excel sheet selection
import io
import logging
import os
from datetime import datetime
from typing import List, Optional

import pandas as pd
from dotenv import load_dotenv
from fastapi import UploadFile, File, HTTPException, APIRouter, BackgroundTasks, Form, Request
from pydantic import BaseModel

from app.routes.delta_routes import DeltaProcessor, get_file_by_id
from app.utils.uuid_generator import generate_uuid

# Load environment variables
load_dotenv()

from app.services.storage_service import uploaded_files

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/files", tags=["files"])


class FileIDsRequest(BaseModel):
    file_ids: List[str]  # Use UUID if you want validation, else use str


# Pydantic models for sheet handling
class SheetInfo(BaseModel):
    sheet_name: str
    row_count: int
    column_count: int
    columns: List[str]


class UpdateSheetRequest(BaseModel):
    sheet_name: str


def detect_leading_zero_columns(content: bytes, filename: str, sheet_name: Optional[str] = None) -> dict:
    """
    Detect columns that contain leading zeros by reading a sample of the file as strings.
    This preserves values like '01', '007', '09' that should stay as strings.
    
    Args:
        content: File content as bytes
        filename: Name of the file
        sheet_name: Sheet name for Excel files
        
    Returns:
        Dictionary mapping column names to dtype ('str' for columns with leading zeros)
    """
    try:
        logger.info("🔍 Detecting columns with leading zeros...")
        
        # Read a small sample as all strings to detect leading zeros
        if filename.lower().endswith('.csv'):
            sample_df = pd.read_csv(
                io.BytesIO(content), 
                dtype=str,  # Read everything as strings
                nrows=100,  # Sample first 100 rows
                encoding='utf-8'
            )
        else:
            sample_df = pd.read_excel(
                io.BytesIO(content),
                sheet_name=sheet_name,
                dtype=str,  # Read everything as strings
                nrows=100,  # Sample first 100 rows
                engine='openpyxl'
            )
        
        dtype_mapping = {}
        leading_zero_columns = []
        
        for col in sample_df.columns:
            has_leading_zeros = False
            
            # Check if any non-null values have leading zeros
            non_null_values = sample_df[col].dropna()
            
            for value in non_null_values.head(20):  # Check first 20 values
                if isinstance(value, str) and value.strip():
                    # Check if it's a numeric string with leading zeros
                    stripped_val = value.strip()
                    
                    # Skip if it contains non-digit characters (except decimal point)
                    if not stripped_val.replace('.', '').replace('-', '').isdigit():
                        continue
                    
                    # Check for leading zeros: starts with 0 and has more than 1 digit
                    if (stripped_val.startswith('0') and 
                        len(stripped_val) > 1 and 
                        stripped_val != '0' and
                        '.' not in stripped_val):  # Don't treat '0.123' as leading zero
                        has_leading_zeros = True
                        break
            
            if has_leading_zeros:
                dtype_mapping[col] = str
                leading_zero_columns.append(col)
                logger.info(f"   📌 Column '{col}' contains leading zeros - will preserve as strings")
        
        logger.info(f"✅ Found {len(leading_zero_columns)} columns with leading zeros: {leading_zero_columns}")
        return dtype_mapping
        
    except Exception as e:
        logger.warning(f"⚠️ Could not detect leading zero columns: {e}")
        return {}


def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean column names by stripping leading/trailing spaces and handling duplicates.
    
    This is the entry point for data cleaning - ensures all column names are properly formatted.
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with cleaned column names
    """
    try:
        original_columns = list(df.columns)
        cleaned_columns = []
        column_mapping = {}
        
        logger.info(f"🧹 Cleaning column names for {len(original_columns)} columns...")
        
        for i, col in enumerate(original_columns):
            if col is None:
                # Handle None column names
                cleaned_col = f"Unnamed_{i}"
                logger.warning(f"  - Found None column at index {i}, renamed to '{cleaned_col}'")
            else:
                # Convert to string and strip spaces
                cleaned_col = str(col).strip()
                
                # Handle empty column names after stripping
                if cleaned_col == "":
                    cleaned_col = f"Unnamed_{i}"
                    logger.warning(f"  - Found empty column at index {i}, renamed to '{cleaned_col}'")
                
                # Log if column was changed
                if str(col) != cleaned_col:
                    logger.info(f"  - Cleaned column: '{col}' → '{cleaned_col}'")
            
            # Handle potential duplicates
            original_cleaned = cleaned_col
            counter = 1
            while cleaned_col in cleaned_columns:
                cleaned_col = f"{original_cleaned}_{counter}"
                counter += 1
                if counter > 1:  # Only log if we actually found duplicates
                    logger.warning(f"  - Duplicate column name detected, renamed to '{cleaned_col}'")
            
            cleaned_columns.append(cleaned_col)
            column_mapping[original_columns[i]] = cleaned_col
        
        # Apply column name changes
        df.columns = cleaned_columns
        
        # Count how many columns were actually changed
        changes_made = sum(1 for orig, clean in zip(original_columns, cleaned_columns) if str(orig) != clean)
        
        if changes_made > 0:
            logger.info(f"✅ Successfully cleaned {changes_made}/{len(original_columns)} column names")
            # Log specific changes for debugging
            for orig, clean in zip(original_columns, cleaned_columns):
                if str(orig) != clean:
                    logger.info(f"   '{orig}' → '{clean}'")
        else:
            logger.info("ℹ️  All column names were already clean")
        
        return df
        
    except Exception as e:
        logger.error(f"Error cleaning column names: {str(e)}")
        # Return original DataFrame if cleaning fails
        return df


def clean_data_values(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean data values by stripping leading/trailing spaces from string columns.
    
    Only processes string/object columns to avoid affecting numeric data.
    Preserves data types and handles NaN values safely.
    
    Args:
        df: Input DataFrame
        
    Returns:
        DataFrame with cleaned string data values
    """
    try:
        logger.info(f"🧼 Cleaning data values for {len(df.columns)} columns...")
        
        cleaned_columns_count = 0
        total_values_cleaned = 0
        
        # Process each column
        for col in df.columns:
            # Only process object/string columns
            if df[col].dtype == 'object':
                # Get non-null values to check if they contain strings
                non_null_values = df[col].dropna()
                if len(non_null_values) == 0:
                    logger.debug(f"  - Skipping empty column '{col}'")
                    continue
                
                # Check if this column contains string data (sample first few values)
                sample_values = non_null_values.head(10)
                string_count = sum(1 for val in sample_values if isinstance(val, str))
                
                # Only clean if at least 50% of sampled values are strings
                if string_count >= len(sample_values) * 0.5:
                    original_values = df[col].copy()
                    
                    # Apply string cleaning only to string values, preserve others
                    def clean_string_value(value):
                        if pd.isna(value):
                            return value  # Keep NaN as-is
                        elif isinstance(value, str):
                            cleaned = value.strip()
                            return cleaned
                        else:
                            return value  # Keep non-strings as-is (numbers, dates, etc.)
                    
                    df[col] = df[col].apply(clean_string_value)
                    
                    # Count how many values were actually changed
                    changes_in_column = 0
                    for orig, clean in zip(original_values, df[col]):
                        if pd.notna(orig) and pd.notna(clean) and str(orig) != str(clean):
                            changes_in_column += 1
                    
                    if changes_in_column > 0:
                        cleaned_columns_count += 1
                        total_values_cleaned += changes_in_column
                        logger.info(f"  - Cleaned column '{col}': {changes_in_column} values trimmed")
                        
                        # Show example of cleaning (first changed value)
                        for orig, clean in zip(original_values, df[col]):
                            if pd.notna(orig) and pd.notna(clean) and str(orig) != str(clean):
                                logger.debug(f"    Example: '{orig}' → '{clean}'")
                                break
                    else:
                        logger.debug(f"  - Column '{col}' already clean (no changes needed)")
                else:
                    logger.debug(f"  - Skipping column '{col}': contains mostly non-string data ({string_count}/{len(sample_values)} strings)")
            else:
                logger.debug(f"  - Skipping numeric/datetime column '{col}' ({df[col].dtype})")
        
        if total_values_cleaned > 0:
            logger.info(f"✅ Successfully cleaned {total_values_cleaned} data values across {cleaned_columns_count} columns")
        else:
            logger.info("ℹ️  All data values were already clean")
        
        return df
        
    except Exception as e:
        logger.error(f"Error cleaning data values: {str(e)}")
        # Return original DataFrame if cleaning fails
        return df


def remove_empty_rows_and_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """
    Remove completely empty rows and columns from DataFrame.
    
    Args:
        df: Input DataFrame
        
    Returns:
        Tuple of (cleaned_df, cleanup_stats)
    """
    try:
        logger.info(f"🗑️  Checking for empty rows and columns in DataFrame ({len(df)} rows, {len(df.columns)} columns)...")
        
        original_shape = df.shape
        cleanup_stats = {
            'original_rows': int(original_shape[0]),
            'original_columns': int(original_shape[1]),
            'removed_rows': 0,
            'removed_columns': 0,
            'empty_column_names': [],
            'final_rows': 0,
            'final_columns': 0
        }
        
        # Remove completely empty columns (all NaN or empty strings)
        empty_columns = []
        for col in df.columns:
            # Check if column is completely empty (all NaN, None, or empty strings after stripping)
            non_empty_values = df[col].dropna()  # Remove NaN/None
            
            if len(non_empty_values) == 0:
                # Column has only NaN/None values
                empty_columns.append(col)
            else:
                # Check if remaining values are all empty strings or whitespace
                string_values = [str(val).strip() for val in non_empty_values if pd.notna(val)]
                non_empty_strings = [val for val in string_values if val != '']
                if len(non_empty_strings) == 0:
                    empty_columns.append(col)
        
        if empty_columns:
            logger.info(f"  - Found {len(empty_columns)} completely empty columns: {empty_columns}")
            df = df.drop(columns=empty_columns)
            cleanup_stats['removed_columns'] = int(len(empty_columns))
            cleanup_stats['empty_column_names'] = list(empty_columns)
        
        # Remove completely empty rows (all NaN or empty strings)
        # Create a mask for non-empty rows
        def is_row_empty(row):
            # Check if all values in the row are empty
            for val in row:
                if pd.notna(val) and str(val).strip() != '':
                    return False
            return True
        
        empty_row_mask = df.apply(is_row_empty, axis=1)
        empty_row_count = empty_row_mask.sum()
        
        if empty_row_count > 0:
            logger.info(f"  - Found {empty_row_count} completely empty rows")
            df = df[~empty_row_mask]
            cleanup_stats['removed_rows'] = int(empty_row_count)
        
        # Reset index after removing rows
        df = df.reset_index(drop=True)
        
        # Update final stats
        cleanup_stats['final_rows'] = int(len(df))
        cleanup_stats['final_columns'] = int(len(df.columns))
        
        # Log results
        if cleanup_stats['removed_rows'] > 0 or cleanup_stats['removed_columns'] > 0:
            logger.info(f"✅ Cleanup completed:")
            logger.info(f"   - Rows: {original_shape[0]} → {len(df)} (removed {cleanup_stats['removed_rows']} empty rows)")
            logger.info(f"   - Columns: {original_shape[1]} → {len(df.columns)} (removed {cleanup_stats['removed_columns']} empty columns)")
        else:
            logger.info("ℹ️  No empty rows or columns found")
        
        return df, cleanup_stats
        
    except Exception as e:
        logger.error(f"Error removing empty rows/columns: {str(e)}")
        # Return original DataFrame and empty stats if cleaning fails
        return df, {
            'original_rows': int(len(df)),
            'original_columns': int(len(df.columns)),
            'removed_rows': 0,
            'removed_columns': 0,
            'empty_column_names': [],
            'final_rows': int(len(df)),
            'final_columns': int(len(df.columns))
        }


def preserve_integer_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert float columns back to integers where all values are whole numbers.
    This prevents 15 from being displayed as 15.0 in reconciliation and other processes.
    """
    try:
        logger.info(f"🔢 Preserving integer types in {len(df.columns)} columns...")
        converted_columns = []
        
        for col in df.columns:
            if df[col].dtype == 'float64':
                # Check if all non-null values are whole numbers
                non_null_values = df[col].dropna()
                if len(non_null_values) > 0:
                    try:
                        # Check if all values are integers (no decimal part)
                        is_integer_column = all(
                            float(val).is_integer() for val in non_null_values 
                            if pd.notna(val) and isinstance(val, (int, float))
                        )
                        
                        if is_integer_column:
                            # Convert to Int64 (pandas nullable integer type) to handle NaN values
                            df[col] = df[col].astype('Int64')
                            converted_columns.append(col)
                            logger.debug(f"  - Converted column '{col}' from float64 to Int64")
                    except (ValueError, TypeError):
                        # Skip columns that can't be converted
                        continue
        
        if converted_columns:
            logger.info(f"✅ Preserved integer types in {len(converted_columns)} columns: {converted_columns[:5]}{'...' if len(converted_columns) > 5 else ''}")
        else:
            logger.info("ℹ️  No float columns needed integer type preservation")
                        
        return df
    except Exception as e:
        logger.warning(f"Warning: Could not preserve integer types: {str(e)}")
        return df


def normalize_datetime_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize datetime columns to consistent YYYY-MM-DD string format.
    Uses the comprehensive date parsing from shared date utilities.

    Args:
        df: Input DataFrame

    Returns:
        DataFrame with normalized date columns
    """
    try:
        from app.utils.date_utils import normalize_date_value
        
        # Aggressive date detection - check ALL columns for date content
        all_columns = list(df.columns)
        datetime_columns_detected = df.select_dtypes(include=['datetime64[ns]']).columns
        converted_date_columns = []
        
        logger.info(f"🔍 Checking all {len(all_columns)} columns for date content...")
        if len(datetime_columns_detected) > 0:
            logger.info(f"  - Pandas auto-detected {len(datetime_columns_detected)} datetime columns: {list(datetime_columns_detected)}")

        for col in all_columns:
            # Sample some non-null values to check if they look like dates
            non_null_values = df[col].dropna()
            if len(non_null_values) == 0:
                logger.debug(f"  - Skipping empty column '{col}'")
                continue

            # Test a sample of values using the robust date parser
            sample_size = min(20, len(non_null_values))  # Check up to 20 values
            sample_values = non_null_values.head(sample_size).tolist()

            date_like_count = 0
            for value in sample_values:
                # Test all types of values (strings, numbers, datetime objects)
                parsed_date_str = normalize_date_value(value)
                if parsed_date_str is not None:
                    date_like_count += 1

            # Use 70% threshold for conservative date detection
            # This prevents false positives on mixed numeric/ID columns
            detection_threshold = 0.7
            if date_like_count >= sample_size * detection_threshold:
                try:
                    original_dtype = str(df[col].dtype)
                    logger.info(f"📅 Converting column '{col}' to normalized dates ({date_like_count}/{sample_size} samples are date-like, type: {original_dtype})")

                    # Apply the robust date parser to the entire column
                    def convert_to_date_string(value):
                        if pd.isna(value):
                            return None
                        parsed_date_str = normalize_date_value(value)
                        if parsed_date_str is not None:
                            return parsed_date_str  # Already in YYYY-MM-DD format
                        return str(value)  # Convert to string if not parseable as date
                    
                    df[col] = df[col].apply(convert_to_date_string)
                    converted_date_columns.append(col)
                    logger.info(f"  ✅ Successfully converted '{col}' from {original_dtype} to YYYY-MM-DD strings")

                except Exception as e:
                    logger.warning(f"  ❌ Failed to convert column '{col}' to dates: {str(e)}")
            else:
                logger.debug(f"  - Skipping '{col}': only {date_like_count}/{sample_size} samples are date-like ({date_like_count/sample_size*100:.1f}%)")

        if converted_date_columns:
            logger.info(f"🎉 Successfully normalized {len(converted_date_columns)} columns to YYYY-MM-DD format: {converted_date_columns}")
        else:
            logger.info("ℹ️  No date columns detected for normalization")
        
        # Final validation - check if any datetime columns still exist
        remaining_datetime_cols = df.select_dtypes(include=['datetime64[ns]']).columns
        if len(remaining_datetime_cols) > 0:
            logger.warning(f"⚠️  WARNING: {len(remaining_datetime_cols)} datetime columns still exist after normalization: {list(remaining_datetime_cols)}")
            # Force convert any remaining datetime columns
            for col in remaining_datetime_cols:
                df[col] = df[col].apply(lambda x: normalize_date_value(x) if pd.notna(x) else None)
                logger.info(f"  🔧 Force-converted remaining datetime column: {col}")
        else:
            logger.info("✅ SUCCESS: All datetime objects converted to YYYY-MM-DD strings")

        return df

    except Exception as e:
        logger.error(f"Error normalizing datetime columns: {str(e)}")
        # Return original DataFrame if normalization fails
        return df


def extract_excel_sheet_info(content: bytes, filename: str) -> List[SheetInfo]:
    """Extract basic information about all sheets in an Excel file (lightweight operation)"""
    try:
        # Read the Excel file to get all sheet names (no data processing)
        excel_file = pd.ExcelFile(io.BytesIO(content))
        sheet_info = []

        for sheet_name in excel_file.sheet_names:
            try:
                # Read only the first few rows to get basic metadata (fast operation)
                df_sample = pd.read_excel(
                    excel_file, 
                    sheet_name=sheet_name, 
                    nrows=0  # Read only headers, no data rows
                )
                
                # Get total row count without reading all data (if possible)
                try:
                    # Try to get sheet dimensions from openpyxl for better performance
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                    ws = wb[sheet_name]
                    row_count = ws.max_row - 1  # Subtract header row
                    wb.close()
                except:
                    # Fallback: read all data to get row count (slower)
                    df_full = pd.read_excel(excel_file, sheet_name=sheet_name)
                    row_count = len(df_full)
                
                sheet_info.append(SheetInfo(
                    sheet_name=sheet_name,
                    row_count=max(0, row_count),  # Ensure non-negative
                    column_count=len(df_sample.columns),
                    columns=df_sample.columns.tolist()
                ))
                
            except Exception as e:
                logger.warning(f"Could not read sheet '{sheet_name}' in {filename}: {str(e)}")
                # Add sheet with minimal info if we can't read it
                sheet_info.append(SheetInfo(
                    sheet_name=sheet_name,
                    row_count=0,
                    column_count=0,
                    columns=[]
                ))

        return sheet_info
    except Exception as e:
        logger.error(f"Error extracting sheet information from {filename}: {str(e)}")
        return []


@router.post("/analyze-sheets")
async def analyze_excel_sheets(file: UploadFile = File(...)):
    """
    Analyze Excel file to get available sheets without uploading
    """
    try:
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(400, "Only Excel files are supported for sheet analysis")

        # Read content
        content = await file.read()

        # Extract sheet information
        sheet_info = extract_excel_sheet_info(content, file.filename)

        return {
            "success": True,
            "sheets": [sheet.dict() for sheet in sheet_info],
            "message": f"Found {len(sheet_info)} sheets"
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error analyzing Excel sheets: {e}")
        return {
            "success": False,
            "error": str(e),
            "sheets": []
        }


@router.post("/validate-name")
async def validate_file_name(request: dict):
    """
    Validate if a filename is available
    """
    try:
        filename = request.get('filename', '').strip()

        if not filename:
            return {
                "isValid": False,
                "error": "Filename is required"
            }

        # Check if name already exists
        for file_id, file_data in uploaded_files.items():
            existing_name = file_data["info"].get("custom_name") or file_data["info"]["filename"]
            if existing_name.lower() == filename.lower():
                return {
                    "isValid": False,
                    "error": "A file with this name already exists"
                }

        return {
            "isValid": True,
            "message": "Filename is available"
        }

    except Exception as e:
        logger.error(f"Error validating filename: {e}")
        return {
            "isValid": False,
            "error": "Validation failed"
        }


@router.post("/upload")
async def upload_file(
        background_tasks: BackgroundTasks,
        file: UploadFile = File(...),
        sheet_name: str = Form(None),
        custom_name: str = Form(None)
):
    """
    Upload file with optional sheet selection and custom naming
    Enhanced with Excel sheet selection and file naming
    """
    try:
        if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(400, "Only CSV and Excel files are supported")

        # Get max file size from environment
        max_file_size = int(os.getenv("MAX_FILE_SIZE", "100")) * 1024 * 1024

        file_id = generate_uuid('file')

        # Log file upload start
        logger.info(f"Starting upload of file: {file.filename}")

        # Read content in chunks for large files
        content = await file.read()
        file_size = len(content)

        # Check file size limit
        if file_size > max_file_size:
            raise HTTPException(413, f"File too large. Maximum size: {max_file_size / (1024 * 1024):.1f}MB")

        logger.info(f"File size: {file_size / (1024 * 1024):.2f}MB")

        # Initialize variables
        is_excel = file.filename.lower().endswith(('.xlsx', '.xls'))
        df = None

        # Process file with leading zero preservation
        try:
            # Step 1: Detect columns with leading zeros first
            dtype_mapping = detect_leading_zero_columns(content, file.filename, sheet_name)
            
            if file.filename.lower().endswith('.csv'):
                # Use optimized CSV reading with leading zero preservation
                df = pd.read_csv(
                    io.BytesIO(content),
                    low_memory=False,  # Don't infer dtypes chunk by chunk
                    encoding='utf-8',
                    dtype=dtype_mapping if dtype_mapping else None  # Preserve leading zero columns as strings
                )
            else:
                # For Excel files, use specified sheet or default with leading zero preservation
                if sheet_name:
                    df = pd.read_excel(
                        io.BytesIO(content),
                        sheet_name=sheet_name,
                        engine='openpyxl',
                        dtype=dtype_mapping if dtype_mapping else None  # Preserve leading zero columns as strings
                        # if file.filename.lower().endswith('.xlsx') else 'xlrd'
                    )
                else:
                    # Use first sheet if no sheet specified
                    df = pd.read_excel(
                        io.BytesIO(content),
                        engine='openpyxl',
                        dtype=dtype_mapping if dtype_mapping else None  # Preserve leading zero columns as strings
                        # if file.filename.lower().endswith('.xlsx') else 'xlrd'
                    )

            # Check if we should use fast parallel cleaning for large files
            use_parallel_cleaning = len(df) > 50000 or len(df.columns) > 50
            
            if use_parallel_cleaning:
                logger.info(f"🚀 Using parallel cleaning for large dataset ({len(df):,} rows × {len(df.columns)} columns)")
                
                # Import the parallel cleaning module
                from app.utils.parallel_cleaning import clean_dataframe_fast
                
                # Use high-performance parallel cleaning
                df, cleanup_stats = clean_dataframe_fast(df, max_workers=None)  # Auto-detect optimal thread count
                
            else:
                logger.info(f"📝 Using standard cleaning for smaller dataset ({len(df):,} rows × {len(df.columns)} columns)")
                
                # Step 1: Remove completely empty rows and columns first
                df, cleanup_stats = remove_empty_rows_and_columns(df)
                
                # Step 2: Clean column names (strip spaces, handle duplicates)
                df = clean_column_names(df)
                
                # Step 3: Clean data values (strip spaces from string data)
                df = clean_data_values(df)
            
            # Step 4: Preserve integer types (prevent 15 -> 15.0 conversion)
            df = preserve_integer_types(df)
            
            # Step 5: Always normalize datetime columns (applies to both paths)
            df = normalize_datetime_columns(df)

        except Exception as e:
            logger.error(f"Error reading file {file.filename}: {str(e)}")
            raise HTTPException(400, f"Error reading file: {str(e)}")


        # Log processing results
        total_rows = len(df)
        total_cols = len(df.columns)
        logger.info(f"Successfully processed {file.filename}: {total_rows:,} rows, {total_cols} columns")

        # Determine final filename
        final_filename = custom_name.strip() if custom_name else file.filename

        # Validate custom name if provided
        if custom_name:
            # Check if custom name already exists
            for existing_file_id, existing_file_data in uploaded_files.items():
                existing_name = existing_file_data["info"].get("custom_name") or existing_file_data["info"]["filename"]
                if existing_name.lower() == custom_name.strip().lower():
                    raise HTTPException(400, f"A file with the name '{custom_name}' already exists")

        file_info = {
            "file_id": file_id,
            "filename": file.filename,  # Original filename
            "custom_name": custom_name.strip() if custom_name else None,  # Custom display name
            "file_type": "excel" if is_excel else "csv",
            "file_size_mb": round(file_size / (1024 * 1024), 2),
            "total_rows": int(total_rows),
            "total_columns": int(total_cols),
            "columns": list(df.columns),
            "upload_time": datetime.utcnow().isoformat(),
            "data_types": {col: str(dtype) for col, dtype in df.dtypes.items()},
            "sheet_name": sheet_name if sheet_name else None,  # Store selected sheet
            "cleanup_performed": cleanup_stats,  # Include data cleaning statistics
        }

        # Store in memory
        uploaded_files[file_id] = {
            "info": file_info,
            "data": df
        }

        # Add cleanup task for very large files (optional)
        cleanup_threshold = int(os.getenv("LARGE_FILE_THRESHOLD", "100000"))
        if total_rows > cleanup_threshold:
            logger.info(f"Large file detected ({total_rows:,} rows). Consider implementing cleanup logic.")

        response_message = f"File uploaded successfully - {total_rows:,} rows processed"
        if sheet_name:
            response_message += f" from sheet '{sheet_name}'"
        if custom_name:
            response_message += f" with custom name '{custom_name}'"
        
        # Enhanced user feedback for data cleaning
        cleanup_warnings = []
        cleanup_details = []
        
        # Check for excessive empty content that might indicate file quality issues
        # Convert numpy types to native Python types to avoid serialization issues
        original_rows = int(cleanup_stats['original_rows'])
        original_columns = int(cleanup_stats['original_columns'])
        removed_rows = int(cleanup_stats['removed_rows'])
        removed_columns = int(cleanup_stats['removed_columns'])
        
        # Calculate percentages of empty content
        empty_row_percentage = (removed_rows / original_rows * 100) if original_rows > 0 else 0
        empty_col_percentage = (removed_columns / original_columns * 100) if original_columns > 0 else 0
        
        # Build cleanup details
        if removed_rows > 0:
            cleanup_details.append(f"{removed_rows} empty rows removed")
        if removed_columns > 0:
            cleanup_details.append(f"{removed_columns} empty columns removed")
            if cleanup_stats.get('empty_column_names'):
                empty_col_names = cleanup_stats['empty_column_names'][:5]  # Show first 5 empty column names
                if len(cleanup_stats['empty_column_names']) > 5:
                    cleanup_details.append(f"Empty columns included: {', '.join(empty_col_names[:3])}, and {len(cleanup_stats['empty_column_names']) - 3} more")
                else:
                    cleanup_details.append(f"Empty columns: {', '.join(empty_col_names)}")
        
        # Add cleanup information to response message
        if cleanup_details:
            response_message += f". ⚡ Data cleanup: {', '.join(cleanup_details)}"
        
        # Generate warnings for excessive empty content
        if empty_row_percentage >= 30:
            cleanup_warnings.append(f"⚠️  High number of empty rows detected ({empty_row_percentage:.1f}% of original file)")
        if empty_col_percentage >= 20:
            cleanup_warnings.append(f"⚠️  High number of empty columns detected ({empty_col_percentage:.1f}% of original file)")
        
        # Add warnings for poor file quality
        if removed_rows > 100 or removed_columns > 10:
            cleanup_warnings.append("💡 Tip: Consider cleaning your Excel/CSV files before upload to improve processing speed")
        
        # Log warnings for monitoring
        if cleanup_warnings:
            for warning in cleanup_warnings:
                logger.warning(f"File quality issue in {final_filename}: {warning}")

        # Build comprehensive cleanup_performed response with parallel processing stats
        cleanup_response = {
            "empty_content_removed": removed_rows > 0 or removed_columns > 0,
            "parallel_processing_used": use_parallel_cleaning,
            "statistics": {
                "original_size": f"{original_rows:,} rows × {original_columns} columns",
                "final_size": f"{total_rows:,} rows × {total_cols} columns",
                "removed_rows": int(removed_rows),
                "removed_columns": int(removed_columns),
                "empty_row_percentage": float(round(empty_row_percentage, 1)),
                "empty_column_percentage": float(round(empty_col_percentage, 1))
            },
            "warnings": cleanup_warnings,
            "details": cleanup_details
        }
        
        # Add parallel processing performance stats if available
        if hasattr(cleanup_stats, 'performance_stats') and cleanup_stats.get('performance_stats'):
            perf_stats = cleanup_stats['performance_stats']
            
            # Add processing time breakdown
            if 'timing' in perf_stats:
                cleanup_response["performance"] = {
                    "total_processing_time": cleanup_stats.get('processing_time_seconds', 0),
                    "timing_breakdown": perf_stats['timing'],
                    "processing_method": "parallel_multi_threaded" if use_parallel_cleaning else "standard_sequential"
                }
                
                # Calculate performance metrics
                total_cells = original_rows * original_columns
                processing_time = cleanup_stats.get('processing_time_seconds', 0)
                
                if processing_time > 0:
                    cells_per_second = int(total_cells / processing_time)
                    cleanup_response["performance"]["cells_per_second"] = cells_per_second
                    cleanup_response["performance"]["megacells_per_second"] = round(cells_per_second / 1000000, 2)
                    
            # Add additional parallel-specific stats
            if use_parallel_cleaning:
                cleanup_response["parallel_stats"] = {
                    "max_workers": cleanup_stats.get('performance_stats', {}).get('max_workers', 'auto'),
                    "cleaned_values": cleanup_stats.get('cleaned_values', 0),
                    "normalized_date_columns": cleanup_stats.get('normalized_date_columns', [])
                }
        else:
            # Standard processing stats
            if use_parallel_cleaning:
                cleanup_response["performance"] = {
                    "processing_method": "parallel_multi_threaded",
                    "note": "Parallel processing used for improved performance"
                }
            else:
                cleanup_response["performance"] = {
                    "processing_method": "standard_sequential",
                    "note": "Standard processing used for smaller dataset"
                }

        return {
            "success": True,
            "message": response_message,
            "data": file_info,
            "cleanup_performed": cleanup_response
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload error for {file.filename}: {e}")
        raise HTTPException(500, f"Upload failed: {str(e)}")


@router.get("/{file_id}/sheets")
async def get_file_sheets(file_id: str):
    """Get available sheets for an Excel file"""
    if file_id not in uploaded_files:
        raise HTTPException(404, "File not found")

    file_data = uploaded_files[file_id]
    file_info = file_data["info"]

    if not file_info.get("is_excel", False):
        raise HTTPException(400, "File is not an Excel file")

    return {
        "success": True,
        "data": {
            "file_id": file_id,
            "filename": file_info["filename"],
            "current_sheet": file_info.get("selected_sheet"),
            "available_sheets": file_info.get("available_sheets", [])
        }
    }


@router.post("/{file_id}/select-sheet")
async def select_sheet(file_id: str, request: UpdateSheetRequest):
    """Switch to a different sheet in an Excel file"""
    if file_id not in uploaded_files:
        raise HTTPException(404, "File not found")

    file_data = uploaded_files[file_id]
    file_info = file_data["info"]

    if not file_info.get("is_excel", False):
        raise HTTPException(400, "File is not an Excel file")

    # Check if requested sheet exists
    available_sheets = [sheet["sheet_name"] for sheet in file_info.get("available_sheets", [])]
    if request.sheet_name not in available_sheets:
        raise HTTPException(400, f"Sheet '{request.sheet_name}' not found. Available sheets: {available_sheets}")

    try:
        # Load the new sheet with leading zero preservation
        raw_content = file_data["raw_content"]
        dtype_mapping = detect_leading_zero_columns(raw_content, file_info["filename"], request.sheet_name)
        df = pd.read_excel(
            io.BytesIO(raw_content),
            sheet_name=request.sheet_name,
            engine='openpyxl' if file_info["filename"].lower().endswith('.xlsx') else 'xlrd',
            dtype=dtype_mapping if dtype_mapping else None
        )

        # Apply the same cleaning pipeline as file upload
        # Step 1: Remove completely empty rows and columns
        df, cleanup_stats = remove_empty_rows_and_columns(df)
        
        # Step 2: Clean column names for the new sheet
        df = clean_column_names(df)
        
        # Step 3: Clean data values
        df = clean_data_values(df)
        
        # Step 4: Normalize datetime columns for the new sheet
        df = normalize_datetime_columns(df)
        
        # Update stored data
        uploaded_files[file_id]["data"] = df
        uploaded_files[file_id]["info"]["selected_sheet"] = request.sheet_name
        uploaded_files[file_id]["info"]["total_rows"] = len(df)
        uploaded_files[file_id]["info"]["total_columns"] = len(df.columns)
        uploaded_files[file_id]["info"]["columns"] = list(df.columns)
        uploaded_files[file_id]["info"]["data_types"] = {col: str(dtype) for col, dtype in df.dtypes.items()}

        logger.info(
            f"Switched to sheet '{request.sheet_name}' for file {file_info['filename']}: {len(df):,} rows, {len(df.columns)} columns")

        # Create response message with cleanup information
        response_message = f"Successfully switched to sheet '{request.sheet_name}'"
        if cleanup_stats['removed_rows'] > 0 or cleanup_stats['removed_columns'] > 0:
            cleanup_details = []
            if cleanup_stats['removed_rows'] > 0:
                cleanup_details.append(f"{cleanup_stats['removed_rows']} empty rows removed")
            if cleanup_stats['removed_columns'] > 0:
                cleanup_details.append(f"{cleanup_stats['removed_columns']} empty columns removed")
            response_message += f". Data cleanup: {', '.join(cleanup_details)}"

        return {
            "success": True,
            "message": response_message,
            "data": {
                "selected_sheet": request.sheet_name,
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "columns": list(df.columns),
                "cleanup_performed": cleanup_stats  # Include cleanup stats
            }
        }

    except Exception as e:
        logger.error(f"Error switching to sheet '{request.sheet_name}': {str(e)}")
        raise HTTPException(500, f"Failed to switch sheet: {str(e)}")


@router.get("/")
async def list_files():
    """List all uploaded files with enhanced information"""
    try:
        files = []
        total_rows = 0
        total_size_mb = 0

        for file_id, file_data in uploaded_files.items():
            file_info = file_data["info"].copy()
            files.append(file_info)
            total_rows += file_info.get("total_rows", 0)
            total_size_mb += file_info.get("file_size_mb", 0)

        return {
            "success": True,
            "message": f"Retrieved {len(files)} files",
            "data": {
                "files": files,
                "summary": {
                    "total_files": len(files),
                    "total_rows": total_rows,
                    "total_size_mb": round(total_size_mb, 2)
                }
            }
        }
    except Exception as e:
        logger.error(f"List files error: {e}")
        return {
            "success": False,
            "message": f"Failed to list files: {str(e)}",
            "data": {"files": [], "total_count": 0}
        }


@router.get("/{file_id}")
async def get_file_info(file_id: str, include_sample: bool = False, sample_rows: int = 10):
    """Get detailed information about a specific file"""
    if file_id not in uploaded_files:
        raise HTTPException(404, "File not found")

    file_data = uploaded_files[file_id]
    df = file_data["data"]

    response_data = {
        "info": file_data["info"],
    }

    if include_sample:
        # Get sample data (limited for large files)
        max_sample_rows = int(os.getenv("MAX_SAMPLE_ROWS", "100"))
        sample_data = df.head(min(sample_rows, max_sample_rows)).to_dict(orient='records')
        response_data["sample_data"] = sample_data

        # Enhanced column statistics
        column_stats = {}
        for col in df.columns:
            col_data = df[col]
            column_stats[col] = {
                "dtype": str(col_data.dtype),
                "null_count": int(col_data.isna().sum()),
                "null_percentage": round((col_data.isna().sum() / len(df)) * 100, 2),
                "unique_count": int(col_data.nunique()),
                "sample_values": col_data.dropna().head(5).astype(str).tolist()
            }

            # Add numeric statistics for numeric columns
            if pd.api.types.is_numeric_dtype(col_data):
                column_stats[col].update({
                    "min": float(col_data.min()) if not col_data.isna().all() else None,
                    "max": float(col_data.max()) if not col_data.isna().all() else None,
                    "mean": float(col_data.mean()) if not col_data.isna().all() else None
                })

        response_data["column_statistics"] = column_stats

    return {
        "success": True,
        "message": "File details retrieved",
        "data": response_data
    }


@router.delete("/{file_id}")
async def delete_file(file_id: str):
    """Delete a file from memory"""
    if file_id not in uploaded_files:
        raise HTTPException(404, "File not found")

    file_info = uploaded_files[file_id]["info"]
    del uploaded_files[file_id]

    logger.info(f"Deleted file: {file_info['filename']} ({file_info['total_rows']:,} rows)")

    return {
        "success": True,
        "message": f"File {file_info['filename']} deleted successfully"
    }


@router.post("/bulk-delete")
async def bulk_delete(request: FileIDsRequest):
    file_ids = request.file_ids

    for file_id in file_ids:
        await delete_file(file_id)


@router.get("/{file_id}/preview")
async def preview_file_data(
        file_id: str,
        start_row: int = 0,
        num_rows: int = 100,
        columns: str = None
):
    """Get a preview of file data with pagination"""
    if file_id not in uploaded_files:
        raise HTTPException(404, "File not found")

    df = uploaded_files[file_id]["data"]

    # Limit preview rows based on environment variable
    max_preview_rows = int(os.getenv("MAX_PREVIEW_ROWS", "1000"))
    num_rows = min(num_rows, max_preview_rows)

    # Parse columns if specified
    if columns:
        requested_cols = [col.strip() for col in columns.split(',')]
        available_cols = [col for col in requested_cols if col in df.columns]
        if available_cols:
            df = df[available_cols]

    # Apply pagination
    end_row = min(start_row + num_rows, len(df))
    preview_df = df.iloc[start_row:end_row]

    return {
        "success": True,
        "data": {
            "rows": preview_df.to_dict(orient='records'),
            "pagination": {
                "start_row": start_row,
                "end_row": end_row,
                "total_rows": len(df),
                "returned_rows": len(preview_df)
            }
        }
    }


@router.get("/{file_id}/columns/{column_name}/unique-values")
async def get_column_unique_values(
        file_id: str,
        column_name: str,
        limit: Optional[int] = 1000,
        request: Request = None
):
    """Get unique values for a specific column in a file with cascading filter support"""

    try:
        # Get the file DataFrame
        df = await get_file_by_id(file_id)

        if column_name not in df.columns:
            raise HTTPException(
                status_code=404,
                detail=f"Column '{column_name}' not found in file"
            )

        # Apply cascading filters from query parameters
        filtered_df = df.copy()
        applied_filters = []
        
        if request:
            # Parse filter parameters (format: filter_columnname=value1,value2)
            for param_name, param_value in request.query_params.items():
                if param_name.startswith('filter_') and param_name != f'filter_{column_name}':
                    filter_column = param_name[7:]  # Remove 'filter_' prefix
                    
                    if filter_column in filtered_df.columns:
                        # Parse comma-separated filter values
                        filter_values = [v.strip() for v in param_value.split(',') if v.strip()]
                        
                        if filter_values:
                            # Apply filter: keep only rows where the filter column has one of the specified values
                            filtered_df = filtered_df[filtered_df[filter_column].astype(str).isin(filter_values)]
                            applied_filters.append({
                                'column': filter_column,
                                'values': filter_values,
                                'matched_rows': len(filtered_df)
                            })
                            
                            logger.info(f"Applied cascading filter on {filter_column}: {filter_values} -> {len(filtered_df)} rows remaining")
        
        # Get the column data from filtered DataFrame
        column_data = filtered_df[column_name].dropna()

        if len(column_data) == 0:
            return {
                "file_id": file_id,
                "column_name": column_name,
                "unique_values": [],
                "total_unique": 0,
                "is_date_column": False,
                "sample_values": []
            }

        # Check if this might be a date column by sampling some values using shared date utilities
        from app.utils.date_utils import normalize_date_value
        sample_size = min(50, len(column_data))
        sample_values = column_data.sample(n=sample_size).tolist()

        # Test if this looks like a date column using updated date utilities
        parsed_dates = 0
        for value in sample_values[:10]:  # Test first 10 samples
            if normalize_date_value(value) is not None:
                parsed_dates += 1

        is_date_column = parsed_dates >= 5  # If 5+ out of 10 samples parse as dates

        # Get unique values
        unique_values = column_data.unique()

        # Limit the number of unique values returned
        if len(unique_values) > limit:
            unique_values = unique_values[:limit]

        # Process values based on whether it's a date column
        processed_values = []

        for value in unique_values:
            if pd.isna(value):
                continue

            if is_date_column:
                # Try to parse as date using shared date utilities
                parsed_date_str = normalize_date_value(value)
                if parsed_date_str is not None:
                    # Already in YYYY-MM-DD format from shared utilities
                    processed_values.append({
                        "original_value": value,
                        "display_value": parsed_date_str,
                        "sort_value": parsed_date_str,
                        "is_date": True
                    })
                else:
                    # Not a valid date, treat as string
                    processed_values.append({
                        "original_value": value,
                        "display_value": str(value),
                        "sort_value": str(value).lower(),
                        "is_date": False
                    })
            else:
                # Regular string/numeric value
                processed_values.append({
                    "original_value": value,
                    "display_value": str(value),
                    "sort_value": str(value).lower() if not pd.api.types.is_numeric_dtype(type(value)) else value,
                    "is_date": False
                })

        # Sort the processed values
        if is_date_column:
            # Sort dates chronologically
            processed_values.sort(key=lambda x: x["sort_value"])
        else:
            # Sort alphabetically/numerically
            processed_values.sort(key=lambda x: str(x["sort_value"]))

        # Extract just the display values for the response
        unique_display_values = [item["display_value"] for item in processed_values]

        # Get some sample values for debugging
        sample_display_values = unique_display_values[:10]

        return {
            "file_id": file_id,
            "column_name": column_name,
            "unique_values": unique_display_values,
            "total_unique": len(column_data.unique()),
            "returned_count": len(unique_display_values),
            "is_date_column": is_date_column,
            "sample_values": sample_display_values,
            "has_more": len(column_data.unique()) > limit,
            "applied_filters": applied_filters,
            "filtered_from_rows": len(df) if applied_filters else None,
            "filtered_to_rows": len(filtered_df) if applied_filters else None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting unique values for column {column_name} in file {file_id}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Failed to get unique values: {str(e)}"
        )
